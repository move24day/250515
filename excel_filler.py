# excel_filler.py

import openpyxl
import io
import streamlit as st
import os
import traceback
from datetime import date
import re
import utils

try:
    import data
except ImportError:
    st.error("data.py 파일을 찾을 수 없습니다. excel_filler.py와 같은 폴더에 있는지 확인하세요.")
    data = None

def get_tv_qty(state_data):
    if not data or not hasattr(data, 'items') or not isinstance(data.items, dict):
        return 0
    total_tv_qty = 0
    tv_keys = [key for key in data.items if key.startswith("TV(")]
    for tv_item_name in tv_keys:
        total_tv_qty += utils.get_item_qty(state_data, tv_item_name)
    return total_tv_qty

def get_method_label_prefix(method_string):
    """ "사다리차 🪜" -> "사다리", "스카이 🏗️" -> "스카이" """
    if not method_string or not isinstance(method_string, str):
        return ""
    return method_string.split(" ")[0]


def fill_final_excel_template(state_data, calculated_cost_items, total_cost_overall, personnel_info):
    if not data:
        st.error("data.py 모듈 로드 실패로 Excel 생성을 진행할 수 없습니다.")
        return None

    try:
        script_dir = os.path.dirname(__file__) if "__file__" in locals() else "."
        final_xlsx_path = os.path.join(script_dir, "final.xlsx")

        if not os.path.exists(final_xlsx_path):
            st.error(f"템플릿 파일 '{final_xlsx_path}'을 찾을 수 없습니다.")
            return None

        wb = openpyxl.load_workbook(final_xlsx_path)
        ws = wb.active # 일반적으로 'Sheet1' 또는 첫 번째 시트

        # --- 1. 기본 정보 입력 ---
        is_storage = state_data.get('is_storage_move', False)
        is_long_distance = state_data.get('apply_long_distance', False)
        has_via_point = state_data.get('has_via_point', False)

        move_type_parts = []
        if is_storage: move_type_parts.append("보관")
        if has_via_point: move_type_parts.append("경유")
        if is_long_distance: move_type_parts.append("장거리")
        
        base_move_type_from_state = state_data.get('base_move_type', "") # state_data에서 가져옴
        if "사무실" in base_move_type_from_state: move_type_parts.append("사무실")
        elif "가정" in base_move_type_from_state: move_type_parts.append("가정")
        
        move_type_str = " ".join(move_type_parts).strip() or base_move_type_from_state
        ws['J1'] = move_type_str

        ws['C2'] = state_data.get('customer_name', '')
        ws['G2'] = state_data.get('customer_phone', '')
        
        moving_date_val = state_data.get('moving_date')
        if isinstance(moving_date_val, date):
            ws['K3'] = moving_date_val
            ws['K3'].number_format = 'yyyy-mm-dd'
        elif moving_date_val:
            ws['K3'] = str(moving_date_val)
        else:
            ws['K3'] = ''

        ws['C3'] = state_data.get('from_location', '')
        ws['C4'] = state_data.get('to_location', '')

        if has_via_point:
            ws['G4'] = state_data.get('via_point_location', '')
        else:
            ws['G4'] = '' # 값이 없을 때 G4 셀을 비움

        p_info = personnel_info if isinstance(personnel_info, dict) else {}
        try: ws['L5'] = int(p_info.get('final_men', 0) or 0)
        except (ValueError, TypeError): ws['L5'] = 0
        try: ws['L6'] = int(p_info.get('final_women', 0) or 0)
        except (ValueError, TypeError): ws['L6'] = 0

        from_floor_str = str(state_data.get('from_floor', '')).strip()
        ws['D5'] = f"{from_floor_str}층" if from_floor_str else ''
        to_floor_str = str(state_data.get('to_floor', '')).strip()
        ws['D6'] = f"{to_floor_str}층" if to_floor_str else ''
        
        # E5, E6는 출발지/도착지 전체 작업 방법 문자열 (예: "사다리차 🪜")
        ws['E5'] = state_data.get('from_method', '')
        ws['E6'] = state_data.get('to_method', '')
        
        if has_via_point:
            ws['K6'] = state_data.get('via_point_method', '')
        else:
            ws['K6'] = ''

        selected_vehicle = state_data.get('final_selected_vehicle', '')
        vehicle_tonnage = ''
        if isinstance(selected_vehicle, str) and selected_vehicle.strip():
            try:
                match = re.search(r'(\d+(\.\d+)?)', selected_vehicle)
                if match:
                    vehicle_tonnage = match.group(1)
                else:
                    vehicle_tonnage_cleaned = re.sub(r'[^\d.]', '', selected_vehicle)
                    vehicle_tonnage = vehicle_tonnage_cleaned if vehicle_tonnage_cleaned else ''
            except Exception:
                vehicle_tonnage = ''
        elif selected_vehicle: # 숫자가 아닌 다른 타입일 경우 문자열로 변환
             vehicle_tonnage = str(selected_vehicle)
        ws['B7'] = vehicle_tonnage # "톤" 글자 제외하고 숫자만

        dispatched_parts = []
        dispatched_1t = state_data.get('dispatched_1t', 0)
        dispatched_2_5t = state_data.get('dispatched_2_5t', 0)
        dispatched_3_5t = state_data.get('dispatched_3_5t', 0)
        dispatched_5t = state_data.get('dispatched_5t', 0)
        try: dispatched_1t = int(dispatched_1t or 0)
        except: dispatched_1t = 0
        try: dispatched_2_5t = int(dispatched_2_5t or 0)
        except: dispatched_2_5t = 0
        try: dispatched_3_5t = int(dispatched_3_5t or 0)
        except: dispatched_3_5t = 0
        try: dispatched_5t = int(dispatched_5t or 0)
        except: dispatched_5t = 0
            
        if dispatched_1t > 0: dispatched_parts.append(f"1톤: {dispatched_1t}")
        if dispatched_2_5t > 0: dispatched_parts.append(f"2.5톤: {dispatched_2_5t}")
        if dispatched_3_5t > 0: dispatched_parts.append(f"3.5톤: {dispatched_3_5t}")
        if dispatched_5t > 0: dispatched_parts.append(f"5톤: {dispatched_5t}")
        ws['H7'] = ", ".join(dispatched_parts) if dispatched_parts else ''


        # --- 2. 비용 정보 입력 (수정됨) ---
        total_moving_expenses_f22 = 0 # F22에 들어갈 총괄 이사 비용 (VAT, 카드수수료, 작업비, 보관료 제외)
        
        departure_work_cost_f23 = 0
        arrival_work_cost_f24 = 0
        storage_fee_j22 = 0

        # calculated_cost_items를 순회하며 비용 집계
        if calculated_cost_items and isinstance(calculated_cost_items, list):
            for item in calculated_cost_items:
                if isinstance(item, (list, tuple)) and len(item) >= 2:
                    label = str(item[0])
                    try:
                        amount = int(item[1] or 0)
                    except (ValueError, TypeError):
                        amount = 0

                    # F22 총괄 이사 비용에 포함될 항목들
                    if label == '기본 운임':
                        total_moving_expenses_f22 += amount
                    elif label == '날짜 할증':
                        total_moving_expenses_f22 += amount
                    elif "조정 금액" in label: # "할증 조정 금액", "할인 조정 금액"
                        total_moving_expenses_f22 += amount
                    elif label == '장거리 운송료':
                        total_moving_expenses_f22 += amount
                    elif label == '폐기물 처리' or label == '폐기물 처리(톤)': # calculations.py의 레이블 확인 필요
                        total_moving_expenses_f22 += amount
                    elif label == '추가 인력':
                        total_moving_expenses_f22 += amount
                    elif label == '지방 사다리 추가요금':
                        total_moving_expenses_f22 += amount
                    elif label == '경유지 추가요금':
                        total_moving_expenses_f22 += amount
                    
                    # F23, F24 작업 비용 항목들
                    elif label == '출발지 사다리차' or label == '출발지 스카이 장비':
                        departure_work_cost_f23 = amount
                    elif label == '도착지 사다리차' or label == '도착지 스카이 장비':
                        arrival_work_cost_f24 = amount
                    
                    # J22 보관료 항목
                    elif label == '보관료':
                        storage_fee_j22 = amount
                    
                    # VAT 및 카드 수수료는 total_cost_overall에 이미 반영되어 있으므로 F25에서 사용, 여기서는 합산 안함

        # F22 셀: 총괄 이사 비용 (작업비, 보관료, VAT, 카드수수료 제외)
        ws['F22'] = total_moving_expenses_f22

        # B23, F23 셀: 출발지 작업 방식 및 비용
        from_method_str = state_data.get('from_method', '')
        ws['B23'] = "출발" + get_method_label_prefix(from_method_str) if from_method_str else "출발작업"
        ws['F23'] = departure_work_cost_f23

        # B24, F24 셀: 도착지 작업 방식 및 비용
        to_method_str = state_data.get('to_method', '')
        ws['B24'] = "도착" + get_method_label_prefix(to_method_str) if to_method_str else "도착작업"
        ws['F24'] = arrival_work_cost_f24
            
        # H22, J22 셀: 보관료
        if storage_fee_j22 > 0 or is_storage: # 보관이사일 경우 항상 레이블 표시, 비용은 있을 때만
            ws['H22'] = "보관료"
            ws['J22'] = storage_fee_j22
        else:
            ws['H22'] = ""
            ws['J22'] = 0
        
        # 계약금, 잔금, 총액 (VAT 및 카드 수수료 포함된 최종 금액 기준)
        deposit_amount_raw = state_data.get('deposit_amount', state_data.get('tab3_deposit_amount', 0))
        try:
            deposit_amount = int(deposit_amount_raw or 0)
        except (ValueError, TypeError):
            deposit_amount = 0
        ws['J23'] = deposit_amount # 계약금

        try: # total_cost_overall은 calculations.py에서 반환된 최종 금액
            total_cost_num_overall = int(total_cost_overall or 0)
        except (ValueError, TypeError):
            total_cost_num_overall = 0
        ws['F25'] = total_cost_num_overall # 총액 (VAT/카드수수료 포함)

        remaining_balance = total_cost_num_overall - deposit_amount
        ws['J24'] = remaining_balance # 잔금


        # --- 3. 고객 요구사항 입력 ---
        special_notes_str = state_data.get('special_notes', '')
        start_row_notes = 26 # B26부터 시작
        max_possible_note_lines = 20 # 예시: 최대 20줄

        # 기존 내용 지우기 (선택적)
        for i in range(max_possible_note_lines):
             clear_cell_addr = f"B{start_row_notes + i}"
             try:
                 if ws[clear_cell_addr].value is not None: # 셀이 존재하고 값이 있을 때만 None으로 설정
                     ws[clear_cell_addr].value = None
             except Exception:
                 pass # 셀이 없거나 접근 불가시 무시

        if special_notes_str:
            notes_parts = [part.strip() for part in special_notes_str.split('.') if part.strip()] # '.' 기준으로 나누고 공백 제거
            for i, part in enumerate(notes_parts):
                if i < max_possible_note_lines: # 최대 줄 수 넘지 않도록
                    target_cell_notes = f"B{start_row_notes + i}"
                    try:
                        ws[target_cell_notes] = part
                    except Exception as e_note:
                         print(f"ERROR [Excel Filler B26+]: Failed to write note to {target_cell_notes}: {e_note}")
        else: # 고객 요구사항이 없을 경우, 첫 줄만 비움 (이미 위에서 처리됨)
             try:
                 if ws['B26'].value is not None: ws['B26'] = None
             except Exception: pass


        # --- 4. 품목 수량 입력 ---
        original_jangrong_qty = utils.get_item_qty(state_data, '장롱')
        jangrong_formatted_qty = "0.0"
        try:
            calculated_qty = original_jangrong_qty / 3.0
            jangrong_formatted_qty = f"{calculated_qty:.1f}"
        except ZeroDivisionError:
            jangrong_formatted_qty = "0.0"
        except Exception:
            jangrong_formatted_qty = "Error"
        ws['D8'] = jangrong_formatted_qty

        ws['D9'] = utils.get_item_qty(state_data, '더블침대')
        ws['D10'] = utils.get_item_qty(state_data, '서랍장')
        ws['D11'] = utils.get_item_qty(state_data, '서랍장(3단)')
        ws['D12'] = utils.get_item_qty(state_data, '4도어 냉장고')
        ws['D13'] = utils.get_item_qty(state_data, '김치냉장고(일반형)')
        ws['D14'] = utils.get_item_qty(state_data, '김치냉장고(스탠드형)')
        ws['D15'] = utils.get_item_qty(state_data, '소파(3인용)')
        ws['D16'] = utils.get_item_qty(state_data, '소파(1인용)')
        ws['D17'] = utils.get_item_qty(state_data, '식탁(4인)')
        ws['D18'] = utils.get_item_qty(state_data, '에어컨')
        ws['D19'] = utils.get_item_qty(state_data, '거실장') # "장식장" -> "거실장"으로 변경
        ws['D20'] = utils.get_item_qty(state_data, '피아노(디지털)')
        ws['D21'] = utils.get_item_qty(state_data, '세탁기 및 건조기')

        ws['H9'] = utils.get_item_qty(state_data, '사무실책상')
        ws['H10'] = utils.get_item_qty(state_data, '책상&의자')
        ws['H11'] = utils.get_item_qty(state_data, '책장')
        # ws['H12']에 "오디오 및 스피커" 또는 "컴퓨터&모니터"가 있었다면 해당 라인 수정 또는 추가
        # 예시: 엑셀 템플릿의 H12 셀이 "컴퓨터/모니터" 수량 칸이라면
        ws['H12'] = utils.get_item_qty(state_data, '컴퓨터&모니터') # "오디오 및 스피커" 대신 사용

        ws['H15'] = utils.get_item_qty(state_data, '바구니')
        ws['H16'] = utils.get_item_qty(state_data, '중박스')
        ws['H19'] = utils.get_item_qty(state_data, '화분')
        ws['H20'] = utils.get_item_qty(state_data, '책바구니')

        ws['L8'] = utils.get_item_qty(state_data, '스타일러')
        ws['L9'] = utils.get_item_qty(state_data, '안마기')
        ws['L10'] = utils.get_item_qty(state_data, '피아노(일반)')
        ws['L12'] = get_tv_qty(state_data)
        ws['L16'] = utils.get_item_qty(state_data, '금고')
        ws['L17'] = utils.get_item_qty(state_data, '앵글')


        # --- 5. 완료된 엑셀 파일을 메모리에 저장 ---
        output = io.BytesIO()
        wb.save(output)
        output.seek(0) # 버퍼 포인터 리셋
        return output.getvalue() # 바이트 데이터 반환

    except FileNotFoundError:
        st.error(f"Excel 템플릿 파일 '{final_xlsx_path}'을(를) 찾을 수 없습니다.")
        return None
    except Exception as e:
        st.error(f"Excel 생성 중 오류 발생: {e}")
        traceback.print_exc()
        return None
